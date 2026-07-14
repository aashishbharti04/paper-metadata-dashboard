"""
Research Paper Metadata Extraction System
==========================================
Extracts Paper Title, Author Names and Author Emails from research paper PDFs
and writes them to a single Excel (.xlsx) file — one row per author email,
with the paper info repeated on every row.

Usage:
    python extract_metadata.py <pdf_folder>                          # extract -> Research_Paper_Metadata.xlsx
    python extract_metadata.py <pdf_folder> -o output.xlsx           # custom output file
    python extract_metadata.py <pdf_folder> --append existing.xlsx   # append new papers to an existing sheet
    python extract_metadata.py --check existing.xlsx                 # validate an existing metadata sheet

Requires: PyMuPDF (fitz), openpyxl
"""

import argparse
import os
import re
import sys
import unicodedata

import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = ["Paper ID / Paper Name", "Paper Title", "All Author Names", "Author Email"]

# ---------------------------------------------------------------- email regex
EMAIL_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9._%+\-]*@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
# grouped form: {a, b, c}@domain.com  or  (a; b)@domain.com
GROUP_EMAIL_RE = re.compile(r"[{(\[]([^{}()\[\]@]{1,300})[})\]]\s*@\s*([A-Za-z0-9.\-]+\.[A-Za-z]{2,})")

ABSTRACT_MARKERS = re.compile(
    r"^\W*(abstract|a\s*b\s*s\s*t\s*r\s*a\s*c\s*t|keywords|index terms)\b", re.I
)
# word STEMS with \w* so "University", "Technology", "Laboratories" etc. all match
AFFILIATION_HINTS = re.compile(
    r"\b(universit|institut|college|department|dept|school|facult|laborator|"
    r"academ|hospital|corporat|research|centre|center|gmail|technolog|engineer|"
    r"manager|director|administrat|researcher|consultant|professor|lecturer|"
    r"scientist|analyst|affiliation|location|assistant|associate|student|scholar|"
    r"ltd|inc|llc|pvt|sciences)\w*\b|@|\d{5,}", re.I
)
PLACE_HINTS = re.compile(
    r"\b(india|china|usa|u\.s\.a|united states|united kingdom|america|morocco|"
    r"vietnam|viet nam|sri lanka|malaysia|indonesia|pakistan|bangladesh|nepal|"
    r"saudi arabia|uae|oman|qatar|egypt|nigeria|kenya|ghana|south africa|ireland|"
    r"germany|france|spain|italy|poland|romania|turkey|iran|iraq|jordan|canada|"
    r"australia|japan|korea|singapore|thailand|philippines|brazil|mexico)\b", re.I
)
COMPANY_HINTS = re.compile(
    r"\b(microsoft|google|amazon|ibm|meta|apple|oracle|sap|intel|nvidia|infosys|"
    r"tcs|wipro|accenture|deloitte|capgemini|cognizant|software|solutions|"
    r"technologies|consulting)\b", re.I
)
ACRONYM_PAREN = re.compile(r"\([A-Z][A-Z0-9&.\- ]{2,}\s*\)")  # "(LARMIG)", "(UCLA)"
LABEL_LINE = re.compile(
    r"\b(co[\s\-]?authors?|authors?|name|affiliation|correspondence|location|mail id|email)\s*:", re.I
)
NON_AUTHOR_LINE = re.compile(
    r"\b(proceedings|conference|ieee|springer|elsevier|copyright|doi|issn|isbn|"
    r"vol\.|volume|issue|journal|preprint|arxiv|received|accepted|published|"
    r"corresponding author|www\.|http|abstract)\b", re.I
)


def is_affiliation_text(t):
    return bool(AFFILIATION_HINTS.search(t) or PLACE_HINTS.search(t)
                or COMPANY_HINTS.search(t) or ACRONYM_PAREN.search(t)
                or LABEL_LINE.search(t))


def clean(s: str) -> str:
    s = unicodedata.normalize("NFC", s)
    s = s.replace("­", "")            # soft hyphen
    s = re.sub(r"[ \t ]+", " ", s)
    return s.strip()


def fix_hyphen_wrap(s: str) -> str:
    return re.sub(r"-\s+(?=[a-z])", "", s)


# ---------------------------------------------------------------- extraction
def get_spans(page):
    """Return list of (text, size, y, flags) spans from a page, in reading order."""
    spans = []
    d = page.get_text("dict")
    for block in d.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            line_text = clean("".join(sp["text"] for sp in line.get("spans", [])))
            if not line_text:
                continue
            sizes = [sp["size"] for sp in line["spans"] if sp["text"].strip()]
            size = max(sizes) if sizes else 0
            y = line["bbox"][1]
            spans.append({"text": line_text, "size": round(size, 1), "y": y})
    spans.sort(key=lambda s: s["y"])
    return spans


def extract_title(spans, page_height):
    """Title = the largest-font text in the upper part of page 1 (joined over
    consecutive lines of the same size)."""
    top = [s for s in spans if s["y"] < page_height * 0.55 and len(s["text"]) > 3
           and not NON_AUTHOR_LINE.search(s["text"])
           and not LABEL_LINE.search(s["text"]) and "@" not in s["text"]]
    if not top:
        return "", None
    max_size = max(s["size"] for s in top)
    title_lines, last_y = [], None
    for s in top:
        if abs(s["size"] - max_size) <= 0.5:
            if last_y is not None and s["y"] - last_y > max_size * 3:
                break  # a distant same-size block is not part of the title
            title_lines.append(s)
            last_y = s["y"]
        elif title_lines:
            break
    title = fix_hyphen_wrap(" ".join(t["text"] for t in title_lines))
    title = re.sub(r"\s+", " ", title).strip()
    bottom_y = title_lines[-1]["y"] if title_lines else None
    return title, bottom_y


def looks_like_author_line(text):
    if len(text) > 200 or NON_AUTHOR_LINE.search(text):
        return False
    if EMAIL_RE.search(text) or GROUP_EMAIL_RE.search(text):
        return False
    if is_affiliation_text(text):
        return False
    # authors: words that are mostly capitalized names, commas, 'and', initials
    letters = re.sub(r"[^A-Za-zÀ-ɏ]", "", text)
    if not letters:
        return False
    words = re.findall(r"[A-Za-zÀ-ɏ][A-Za-zÀ-ɏ.'\-]*", text)
    if not words:
        return False
    capped = sum(1 for w in words if w[0].isupper() or w.lower() in ("and", "van", "de", "der", "bin", "al"))
    return capped / len(words) >= 0.7


def split_authors(raw):
    raw = re.sub(r"\[[^\]]*\]", " ", raw)                       # [0000-0003-...] ORCID
    raw = re.sub(r"[¹²³⁴⁵⁶⁷⁸⁹⁰⁺\*†‡§¶#]", " ", raw)             # superscript markers
    raw = re.sub(r"\b(co[\s\-]?authors?|authors?|name)\s*[:\-]\s*", " ", raw, flags=re.I)
    raw = re.sub(r"\bby\b", " ", raw, flags=re.I)
    raw = re.sub(r"\s+", " ", raw)
    parts = re.split(r",|;|\band\b|&", raw, flags=re.I)
    out = []
    for p in parts:
        p = p.strip()
        p = re.sub(r"^[\d\s]+|[\s\d]+$", "", p)
        p = p.strip(" ,;.")
        if (len(p) >= 2 and re.search(r"[A-Za-zÀ-ɏ]{2,}", p)
                and not is_affiliation_text(p)
                and not re.fullmatch(r"(dr|mr|ms|mrs|prof)\.?", p, re.I)):
            out.append(p)
    return out


def extract_authors(spans, title_bottom_y, page_height):
    """Author names usually sit between the title and the abstract/affiliations."""
    if title_bottom_y is None:
        return []
    seen, out = set(), []

    def push(a):
        # merge names wrapped with a hyphen across lines ("Man-" + "ish Kumar")
        if out and out[-1].endswith("-"):
            merged = out[-1][:-1] + a
            seen.discard(out[-1].lower())
            if merged.lower() not in seen:
                seen.add(merged.lower())
                out[-1] = merged
            else:
                out.pop()
            return
        if a.lower() not in seen:
            seen.add(a.lower())
            out.append(a)

    started = False
    for s in spans:
        if s["y"] <= title_bottom_y:
            continue
        if s["y"] > page_height * 0.75:
            break
        text = s["text"]
        if ABSTRACT_MARKERS.match(text):
            break
        # "Author: John Smith" -> keep the name, drop the label
        m = re.match(r"^\W*(co[\s\-]?authors?|authors?|name)\s*:\s*(.+)$", text, re.I)
        if m:
            text = m.group(2)
        if looks_like_author_line(text):
            started = True
            for a in split_authors(text):
                push(a)
        elif started and not is_affiliation_text(text) and len(text) > 80:
            break  # body text started
    # drop unfinished hyphen fragments and strict prefixes of longer names
    cleaned = [a for a in out if not a.endswith("-")]
    return [a for i, a in enumerate(cleaned)
            if not any(j != i and b.lower().startswith(a.lower() + " ")
                       for j, b in enumerate(cleaned))]


def extract_emails(doc, max_pages=2):
    """Emails from the first pages (author block / footnotes). De-duplicated,
    order preserved. Handles {a,b,c}@domain grouped forms."""
    emails, seen = [], set()
    n = min(max_pages, len(doc))
    text = "\n".join(doc[i].get_text() for i in range(n))
    text = clean(text.replace("\n", " \n"))
    # grouped emails first
    for m in GROUP_EMAIL_RE.finditer(text):
        users, domain = m.group(1), m.group(2)
        for u in re.split(r"[,;]", users):
            u = u.strip()
            if re.fullmatch(r"[A-Za-z0-9._%+\-]+", u):
                e = f"{u}@{domain}"
                if e.lower() not in seen:
                    seen.add(e.lower())
                    emails.append(e)
    for m in EMAIL_RE.finditer(text):
        e = m.group(0).rstrip(".")
        if e.lower() not in seen:
            seen.add(e.lower())
            emails.append(e)
    return emails


def process_pdf(path):
    name = os.path.basename(path)
    try:
        doc = fitz.open(path)
    except Exception as exc:
        return {"file": name, "title": "", "authors": [], "emails": [],
                "error": f"cannot open: {exc}"}
    if len(doc) == 0 or not doc[0].get_text().strip():
        doc.close()
        return {"file": name, "title": "", "authors": [], "emails": [],
                "error": "no text layer (scanned PDF? needs OCR)"}
    page = doc[0]
    spans = get_spans(page)
    title, title_y = extract_title(spans, page.rect.height)
    authors = extract_authors(spans, title_y, page.rect.height)
    emails = extract_emails(doc)
    doc.close()
    return {"file": name, "title": title, "authors": authors, "emails": emails, "error": None}


# ---------------------------------------------------------------- excel output
def write_excel(records, out_path, append_to=None):
    if append_to and os.path.exists(append_to):
        wb = load_workbook(append_to)
        ws = wb.active
        existing = {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]}
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Paper Metadata"
        ws.append(HEADERS)
        for c in range(1, 5):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        existing = set()

    added, skipped = 0, []
    for rec in records:
        if rec["file"] in existing:
            skipped.append(rec["file"])
            continue
        authors_joined = ", ".join(rec["authors"])
        rows = rec["emails"] if rec["emails"] else [""]
        for email in rows:
            ws.append([rec["file"], rec["title"], authors_joined, email])
            for c in range(1, 5):
                ws.cell(row=ws.max_row, column=c).font = Font(name="Arial")
            added += 1

    widths = [22, 70, 55, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(out_path)
    return added, skipped


# ---------------------------------------------------------------- check mode
def check_excel(path):
    """Validate a metadata sheet. Returns (n_rows, n_papers, problems)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    problems = []
    if not rows:
        return 0, 0, ["Sheet is empty."]
    header = [str(c).strip() if c else "" for c in rows[0][:4]]
    if header != HEADERS:
        problems.append(f"Header mismatch: {header} (expected {HEADERS})")

    papers = {}
    def cell(row, idx):
        v = row[idx] if len(row) > idx and row[idx] is not None else ""
        return str(v).strip()

    for i, row in enumerate(rows[1:], start=2):
        pid = cell(row, 0)
        title = cell(row, 1)
        authors = cell(row, 2)
        email = cell(row, 3)
        if not pid:
            problems.append(f"Row {i}: missing Paper ID / Name"); continue
        p = papers.setdefault(pid, {"titles": set(), "authors": set(), "emails": [], "rows": []})
        p["titles"].add(title)
        p["authors"].add(authors)
        p["rows"].append(i)
        if email:
            if not EMAIL_RE.fullmatch(email):
                problems.append(f"Row {i}: '{email}' does not look like a valid email")
            if email.lower() in [e.lower() for e in p["emails"]]:
                problems.append(f"Row {i}: duplicate email '{email}' within paper {pid}")
            p["emails"].append(email)
        else:
            problems.append(f"Row {i}: blank Author Email for {pid}")

    for pid, p in papers.items():
        if len(p["titles"]) > 1:
            problems.append(f"{pid}: inconsistent Paper Title across rows {p['rows']}")
        if len(p["authors"]) > 1:
            problems.append(f"{pid}: inconsistent All Author Names across rows {p['rows']}")
        if not p["emails"]:
            problems.append(f"{pid}: no emails at all")

    return len(rows) - 1, len(papers), problems


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Research paper PDF -> Excel metadata extractor")
    ap.add_argument("folder", nargs="?", help="folder containing PDF files")
    ap.add_argument("-o", "--output", default="Research_Paper_Metadata.xlsx")
    ap.add_argument("--append", metavar="XLSX", help="append to an existing metadata Excel")
    ap.add_argument("--check", metavar="XLSX", help="validate an existing metadata Excel and exit")
    args = ap.parse_args()

    if args.check:
        n_rows, n_papers, problems = check_excel(args.check)
        print(f"Checked {n_rows} data rows, {n_papers} papers.")
        if problems:
            print(f"\n{len(problems)} issue(s) found:")
            for pr in problems:
                print("  -", pr)
        else:
            print("No issues found. Sheet is clean.")
        return

    if not args.folder or not os.path.isdir(args.folder):
        ap.error("provide a folder containing PDFs (or use --check)")

    pdfs = sorted(f for f in os.listdir(args.folder) if f.lower().endswith(".pdf"))
    if not pdfs:
        print("No PDF files found in", args.folder); return

    records, warnings = [], []
    for f in pdfs:
        rec = process_pdf(os.path.join(args.folder, f))
        records.append(rec)
        status = []
        if rec["error"]:
            status.append("ERROR: " + rec["error"])
        else:
            if not rec["title"]:
                status.append("no title found")
            if not rec["authors"]:
                status.append("no authors found")
            if not rec["emails"]:
                status.append("no emails found")
        flag = ("  !! " + "; ".join(status)) if status else ""
        print(f"[{f}] title={rec['title'][:60]!r} authors={len(rec['authors'])} emails={len(rec['emails'])}{flag}")
        if status:
            warnings.append(f"{f}: {'; '.join(status)}")

    out = args.append if args.append else args.output
    added, skipped = write_excel(records, out, append_to=args.append)
    print(f"\nWrote {added} rows to {out}")
    if skipped:
        print(f"Skipped (already in sheet): {', '.join(skipped)}")
    if warnings:
        print("\nReview these files manually:")
        for w in warnings:
            print("  -", w)


if __name__ == "__main__":
    main()
